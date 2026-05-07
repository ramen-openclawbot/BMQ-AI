#!/usr/bin/env python3
"""Import reviewed kitchen inventory workbook into the trusted ledger tables.

Default mode validates and previews counts only. Use --apply with Supabase REST
environment variables to stage rows, upsert approved canonical items, and create
idempotent monthly source movements.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import uuid
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime environment check
    raise SystemExit("Missing dependency: openpyxl. Install it before running this script.") from exc

SHEET_NAME = "01_IMPORT_REVIEW"

HEADER_ALIASES = {
    "approval": ["anh duyet", "anh duyet?", "approve", "approval", "decision"],
    "item_code": ["kitchen item code", "item code", "ma hang", "ma item", "ma nvl", "ma ccdc", "code"],
    "month": ["thang", "month", "period", "ky"],
    "name": ["ten chuan ke toan", "ten", "ten hang", "nguyen vat lieu", "item", "item name"],
    "item_type": ["nhom", "loai", "type", "item type"],
    "unit": ["dvt", "don vi", "unit"],
    "standard_unit_cost": ["don gia", "gia chuan", "standard unit cost", "unit cost"],
    "opening_qty": ["ton dau ky", "ton dau", "opening", "opening qty"],
    "purchase_qty": ["nhap trong ky", "nhap", "purchase", "purchase qty"],
    "usage_qty": ["xuat dung", "tieu hao", "usage", "usage qty"],
    "ending_qty": ["kiem ke cuoi ky", "ton cuoi", "ending", "ending qty"],
    "amount": ["thanh tien", "so tien", "amount"],
}


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    text = re.sub(r"[^\w\s-]", " ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_item_type(value: Any) -> str:
    normalized = normalize_text(value)
    if any(token in normalized for token in ("ccdc", "cong cu", "vat tu")):
        return "tool_supply"
    return "ingredient"


def normalize_decision(value: Any) -> str:
    normalized = normalize_text(value).upper()
    if "REJECT" in normalized or "TU CHOI" in normalized or "KHONG DUYET" in normalized:
        return "REJECT"
    if "REVIEW" in normalized or "CAN" in normalized or "XEM LAI" in normalized:
        return "REVIEW"
    if normalized in {"APPROVE", "APPROVED", "DUYET", "DA DUYET", "OK", "YES", "Y"}:
        return "APPROVE"
    return "REVIEW"


def normalize_key(item_type: str, name: str, unit: str) -> str:
    return f"{normalize_item_type(item_type)}:{normalize_text(name)}:{normalize_text(unit)}"


def decimal_or_none(value: Any) -> str | None:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            value = value.replace(",", "").replace(" ", "")
        return str(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return None


def date_or_none(value: Any, fallback: dt.date | None) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().replace(day=1).isoformat()
    if isinstance(value, dt.date):
        return value.replace(day=1).isoformat()
    normalized = normalize_text(value)
    match = re.search(r"(\d{1,2})[/-](\d{4})", normalized)
    if match:
        month, year = int(match.group(1)), int(match.group(2))
        return dt.date(year, month, 1).isoformat()
    return fallback.isoformat() if fallback else None


def source_hash(row: dict[str, Any]) -> str:
    payload = {
        "item_code": row.get("source_item_code"),
        "month": row.get("source_month").isoformat() if hasattr(row.get("source_month"), "isoformat") else row.get("source_month"),
        "name": row["source_item_name"],
        "item_type": row["source_item_type"],
        "unit": row["source_unit"],
        "standard_unit_cost": row["source_standard_unit_cost"],
        "opening_qty": row.get("source_opening_qty"),
        "purchase_qty": row.get("source_purchase_qty"),
        "usage_qty": row.get("source_usage_qty"),
        "ending_qty": row.get("source_ending_qty"),
        "amount": row.get("source_amount"),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()


def load_rows(path: Path, sheet_name: str, fallback_month: dt.date | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Missing required sheet: {sheet_name}. Available: {', '.join(workbook.sheetnames)}")
    sheet = workbook[sheet_name]

    header_row_idx = None
    header_map: dict[str, int] = {}
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        normalized_headers = [normalize_text(cell) for cell in row]
        candidate_map: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in normalized_headers:
                    candidate_map[field] = normalized_headers.index(alias)
                    break
        if "approval" in candidate_map and "name" in candidate_map:
            header_row_idx = row_idx
            header_map = candidate_map
            break

    if header_row_idx is None:
        raise SystemExit("Could not find headers. Required at minimum: Anh duyệt? and item name.")

    parsed_rows: list[dict[str, Any]] = []
    for source_row_number, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        def cell(field: str) -> Any:
            idx = header_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        name = str(cell("name") or "").strip()
        if not name:
            continue

        decision = normalize_decision(cell("approval"))
        unit = str(cell("unit") or "").strip()
        item_type = normalize_item_type(cell("item_type"))
        unit_cost = decimal_or_none(cell("standard_unit_cost"))
        issues = []
        if decision == "APPROVE" and not unit:
            issues.append("missing_unit")
        if decision == "APPROVE" and Decimal(unit_cost or "0") <= 0:
            issues.append("missing_standard_unit_cost")

        parsed_rows.append({
            "source_row_number": source_row_number,
            "source_month": date_or_none(cell("month"), fallback_month),
            "source_item_code": str(cell("item_code") or "").strip() or None,
            "source_item_name": name,
            "source_item_type": item_type,
            "source_unit": unit,
            "source_standard_unit_cost": unit_cost,
            "source_opening_qty": decimal_or_none(cell("opening_qty")),
            "source_purchase_qty": decimal_or_none(cell("purchase_qty")),
            "source_usage_qty": decimal_or_none(cell("usage_qty")),
            "source_ending_qty": decimal_or_none(cell("ending_qty")),
            "source_amount": decimal_or_none(cell("amount")),
            "approval_decision": decision,
            "import_status": "staged" if decision == "APPROVE" and not issues else ("failed" if decision == "APPROVE" else "skipped"),
            "issue_flags": issues,
            "raw_payload": {str(idx + 1): value for idx, value in enumerate(row) if value is not None},
            "source_normalized_key": normalize_key(item_type, name, unit),
        })
        parsed_rows[-1]["source_hash"] = source_hash(parsed_rows[-1])
    return parsed_rows


class SupabaseRest:
    def __init__(self, url: str, key: str) -> None:
        self.base_url = url.rstrip("/") + "/rest/v1"
        self.key = key

    def request(self, method: str, table: str, body: Any | None = None, query: dict[str, str] | None = None, prefer: str | None = None) -> Any:
        encoded_query = urllib.parse.urlencode(query or {})
        url = f"{self.base_url}/{table}" + (f"?{encoded_query}" if encoded_query else "")
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        request = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request) as response:
                text = response.read().decode("utf-8")
                return json.loads(text) if text else None
        except urllib.error.HTTPError as error:
            message = error.read().decode("utf-8")
            raise RuntimeError(f"Supabase REST error {error.code}: {message}") from error

    def select(self, table: str, query: dict[str, str]) -> list[dict[str, Any]]:
        return self.request("GET", table, query=query) or []

    def insert(self, table: str, body: Any, return_representation: bool = True) -> Any:
        prefer = "return=representation" if return_representation else "return=minimal"
        return self.request("POST", table, body=body, prefer=prefer)

    def patch(self, table: str, body: Any, query: dict[str, str]) -> Any:
        return self.request("PATCH", table, body=body, query=query, prefer="return=representation")

    def rpc(self, function_name: str, body: Any) -> Any:
        return self.request("POST", f"rpc/{function_name}", body=body)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--sheet", default=SHEET_NAME)
    parser.add_argument("--period-start", help="Fallback period start, e.g. 2026-03-01")
    parser.add_argument("--period-end", help="Source period end, e.g. 2026-04-30")
    parser.add_argument("--apply", action="store_true", help="Write to Supabase and apply APPROVE rows")
    args = parser.parse_args()

    fallback_month = dt.date.fromisoformat(args.period_start) if args.period_start else None
    rows = load_rows(args.file, args.sheet, fallback_month)
    counts = {
        "APPROVE": sum(1 for row in rows if row["approval_decision"] == "APPROVE"),
        "REVIEW": sum(1 for row in rows if row["approval_decision"] == "REVIEW"),
        "REJECT": sum(1 for row in rows if row["approval_decision"] == "REJECT"),
    }
    print(json.dumps({"rows_total": len(rows), **counts}, ensure_ascii=False, indent=2))

    if not args.apply:
        return 0

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    actor_id = os.environ.get("SUPABASE_IMPORT_ACTOR_ID")
    if not url or not key or not actor_id:
        raise SystemExit("Set SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY, and SUPABASE_IMPORT_ACTOR_ID to apply.")
    try:
        uuid.UUID(actor_id)
    except ValueError as exc:
        raise SystemExit("SUPABASE_IMPORT_ACTOR_ID must be an auth.users UUID for audit attribution.") from exc

    client = SupabaseRest(url, key)
    batch = client.insert("kitchen_inventory_import_batches", {
        "source_file_name": args.file.name,
        "source_sheet_name": args.sheet,
        "source_period_start": args.period_start,
        "source_period_end": args.period_end,
        "status": "previewed",
        "rows_total": len(rows),
        "rows_approved": counts["APPROVE"],
        "rows_review": counts["REVIEW"],
        "rows_rejected": counts["REJECT"],
        "created_by": actor_id,
    })[0]

    staged_rows = [{**row, "batch_id": batch["id"]} for row in rows]
    client.insert("kitchen_inventory_import_rows", staged_rows)
    result = client.rpc("apply_kitchen_inventory_import_batch", {"p_batch_id": batch["id"]})
    print(json.dumps(result, ensure_ascii=False))
    if isinstance(result, dict) and result.get("status") in {"failed", "partial"}:
        raise SystemExit(
            f"Import apply returned {result.get('status')} for batch {batch['id']}: "
            f"applied={result.get('applied', 0)}, skipped={result.get('skipped', 0)}, failed={result.get('failed', 0)}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
